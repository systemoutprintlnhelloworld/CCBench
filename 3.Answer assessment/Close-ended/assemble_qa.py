# 导包
import os

import pandas as pd
from openpyxl import load_workbook

# read_excel方法用于读取excel文件的某一列,返回一个列表


def read_excel(file_path, columnum=0):
    '''
    读取一个excel文件的第一列
    :param file_path: 文件路径
    :param columnum: 列数
    :return: 第一列数据
    '''
    df = pd.read_excel(file_path)
    return df.iloc[:, columnum].values.tolist()

# write_excel方法用于将数据写到指定excel文件的指定位置


def write_excel_1(file_path, data, start_row=1, start_column=1):
    '''
    将数据写到指定excel文件的指定位置
    :param file_path: 文件路径
    :param data: 数据
    :param start_row: 起始行
    :param start_column: 起始列
    :return:
    '''
    # 读取excel文件
    wb = load_workbook(file_path)
    ws = wb.active
    # 写入数据
    for i in range(len(data)):
        for j in range(len(data[i])):
            ws.cell(row=j+start_row, column=i+start_column, value=data[i][j])
            # print(data[i][j])
    # 保存文件
    wb.save(file_path)

# write_excel方法用于将数据写到指定excel文件的指定位置


def write_excel2(file_path, data, start_row=1, start_column=1):
    '''
    将数据写到指定excel文件的指定位置
    :param file_path: 文件路径
    :param data: 数据
    :param start_row: 起始行
    :param start_column: 起始列
    :return:
    '''
    # 读取excel文件
    wb = load_workbook(file_path)
    ws = wb.active
    # 写入数据
    for i in range(len(data)):
        ws.cell(row=i+start_row, column=start_column, value=data[i])
    # 保存文件
    wb.save(file_path)

# read_filename方法用于读取文件夹下指定后缀的文件名


def read_filename(file_dir, suffix, except_files=None):
    '''
    读取文件夹下指定后缀的文件名,这个函数只会读取一层文件夹,不会递归读取
    :param file_dir: 文件夹路径
    :param suffix: 后缀
    :param except_files: 排除的文件列表
    :return: 文件名列表
    '''
    if except_files is None:
        except_files = []
    return [
        file
        for file in os.listdir(file_dir)
        if file.endswith(suffix) and file not in except_files
    ]

# 实现功能:当读取的excel单元格中date内容不为yes/no(不论大小写)时,将其当做NEC(Not Elsewhere Classified)处理


def post_process(data):
    '''
    当读取的excel单元格中date内容不为yes/no(不论大小写)时,将其当做NEC(Not Elsewhere Classified)处理
    :param data: 数据
    :return: 处理后的数据
    '''
    for i, row in enumerate(data):
        for j, cell in enumerate(row):
            data[i][j] = 'NEC' if str(cell).lower() not in [
                'yes', 'no'] else cell
    return data

# 实现功能:仅仅读取excel单元格中内容中的前3个字符中的YES/NO(不论大小写)读入


def read_excel_3char(data):
    '''
    仅仅读取excel单元格中内容中的前3个字符中的YES/NO(不论大小写)读入,并去除其中其他字符,如No,变为NO
    :param data: 数据
    :return: 处理后的数据
    '''
    for i in range(len(data)):
        for j in range(len(data[i])):
            if data[i][j].lower()[:3] == 'yes':
                data[i][j] = 'YES'
                # 其中只要包含No则提取出来,如No,/No./NO?...都识别为No
            elif data[i][j].lower()[:2] == 'no':
                data[i][j] = 'NO'
    return data

# 创建一个函数,用于计算两个单元格的相似度


def similarity_score(cell1, cell2):
    pass


if __name__ == '__main__':
    # 读取D:\研究生\任务3\测试结果\QA\is文件夹下的除了文心一言-格式修正以外所有excel文件
    file_dir = r'D:\研究生\任务3\测试结果\QA\is'
    except_file = ['文心一言-格式修正.xlsx',]
    out_file = r'D:\研究生\任务3\测试结果\is-qa题型结果汇总.xlsx'
    file_list = read_filename(file_dir, '.xlsx', except_file)
    print(file_list)
    # 将每个文件第一列的数据提取出来,汇总到一个excel文件中
    data = []
    for file in file_list:
        file_path = os.path.join(file_dir, file)
        print(file,"file")
        # 如果文件名为bard-revise.xlsx,则读取第一列,否则读第三列
        if file == 'bard-test.xlsx':
            data.append(read_excel(file_path)) 
        # 如果文件名为gpt-4.xlsx,则读取第二列
        elif file == 'gpt-4.xlsx':
          data.append(read_excel(file_path, columnum=1))  
        data.append(read_excel(file_path, columnum=2))
        print(data)
        #$end 
    post_process(data)

    # 写入数据
    write_excel_1(out_file,
                  data, start_row=2, start_column=1)
    # 再读取D:\研究生\任务3\题库\QA\最终版\is\is_qa.xlsx这个文件.将其第三列提取出来,放在最后一列中,其第一行为Ground Truth
    file_path = r'D:\研究生\任务3\题库\VQA\最终题库\is\is_qa.xlsx'
    data = read_excel(file_path, columnum=2)
    # 仅仅读取excel单元格中内容中的前3个字符中的YES/NO(不论大小写)读入
    data = read_excel_3char(data)

    write_excel2(out_file,
                 data, start_row=2, start_column=7)
    print('Done!')
