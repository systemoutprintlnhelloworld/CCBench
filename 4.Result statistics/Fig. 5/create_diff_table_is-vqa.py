import openpyxl

# Load the workbook and select the match sheet
wb = openpyxl.load_workbook(r'D:\\研究生\\任务3\\测试结果\\章节统计-人工\\is-vqa-章节统计.xlsx')
# sheet = wb['match']
sheet = wb.active

# Define the logic to categorize difficulty based on the sum of values from columns A to F
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=5):
    # Calculate the sum of the first six columns (A-e)
    row_sum = sum(cell.value for cell in row if cell.value is not None)

    # Classify difficulty based on the sum: 0-1 = '困难', 2-3 = '普通', 4+ = '简单'
    if row_sum <= 1:
        difficulty = 'hard'
    elif row_sum <= 3:
        difficulty = 'normal'
    else:
        difficulty = 'easy'

    # Write the difficulty into column J (11th column)
    sheet.cell(row=row[0].row, column=19, value=difficulty)

# Create a new sheet to summarize average scores by chapter and difficulty
sheet_diff = wb.create_sheet('难度平均得分表')

# Write the headers for the new summary sheet
headers = ['Model', 'Difficulty', 'Chapter 2',
           'Chapter 4', 'Chapter 5', 'Chapter 6', 'Chapter 7']
for col_num, header in enumerate(headers, start=1):
    sheet_diff.cell(row=1, column=col_num, value=header)

# Models to iterate over
models = ['bard', 'gpt-4', 'llava', 'qwen-vl-max', 'vilt']

# Iterate over models and difficulty to calculate average scores per chapter
for model_idx, model in enumerate(models, start=1):
    for difficulty in ['简单', '普通', '困难']:
        averages = []
        for chapter_num in ['2', '4', '5', '6', '7']:
            total_score = 0
            count = 0
            # Loop through the rows to calculate the average for the current model, chapter, and difficulty
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=20):
                # Ensure that the score is not None and chapter/difficulty match,chapter在第q列,如果这一列的单元格的第一个字符等于chapter_num,就是这一章的数据
                if row[model_idx - 1].value is not None and row[16].value[0] == chapter_num and row[9].value == difficulty:
                    total_score += row[model_idx - 1].value
                    count += 1
            # Properly handle zero values and avoid division errors
            avg_score = total_score / count if count > 0 else 0
            averages.append(avg_score)

        # Write the model and difficulty into the new sheet
        row_position = (model_idx - 1) * 3 + \
            (['简单', '普通', '困难'].index(difficulty) + 1) + 1
        sheet_diff.cell(row=row_position, column=1, value=model)
        sheet_diff.cell(row=row_position, column=2, value=difficulty)

        # Write the calculated averages into the corresponding columns
        for col_num, avg in enumerate(averages, start=3):
            sheet_diff.cell(row=row_position, column=col_num, value=avg)

# Save the updated file
output_file_path = r'D:\\研究生\\任务3\\测试结果\\章节统计-人工\\is-vqa题型结果汇总 - 章节统计-难度表.xlsx'
wb.save(output_file_path)

output_file_path  # Return the path to the updated file
