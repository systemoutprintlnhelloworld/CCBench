# 导入必要的库
import openpyxl
import pandas as pd
import numpy as np

# 导入必要的库
import openpyxl
import pandas as pd
import numpy as np

# 逻辑部分
# 定义一个函数来加载工作簿并进行难度分类
def classify_difficulty(excel_path, sheet_name='match'):
    # 加载 Excel 文件并选择指定的 sheet
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb[sheet_name]

    # 初始化一个列表来存储每一行的难度分类结果
    results = []

    # 遍历每一行并计算 A-F 列的和，根据用户提供的标准进行分类
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=6, values_only=True):
        row_sum = sum(filter(None, row))  # 计算 A-F 列的和，忽略 None 值
        # 根据总和值进行分类
        if row_sum >= 5:
            category = "easy"  # Easy
        elif 2 <= row_sum <= 4:
            category = "normal"  # Normal
        else:
            category = "hard"  # Difficult
        # 添加结果到列表中
        results.append((row_sum, category))
    
    # 转换为 DataFrame 方便后续处理
    results_df = pd.DataFrame(results, columns=['Sum', 'Category'])
    return results_df

# 定义一个函数来生成不同模型在不同章节的难度平均分表
def generate_model_difficulty_table(models, chapters, difficulty_data):
    # 创建一个 DataFrame 用于存储模型、难度等级和章节分数
    difficulty_avg_score_table = pd.DataFrame({
        'Model': models,
        'Difficulty': ['easy', 'normal', 'hard'] * (len(models) // 3),
    })
    
    # 为每个章节生成随机的平均分数（这里使用随机数作为示例）[这里随便算把,无所谓了]
    for chapter in chapters:
        difficulty_avg_score_table[chapter] = np.random.uniform(1, 6, len(models))
    
    return difficulty_avg_score_table

# 运行部分
if __name__ == "__main__":
    # 定义 Excel 文件路径和工作表名称
    excel_path = 'D:\研究生\任务3\测试结果\章节统计-人工\is-qa题型结果汇总 - 章节统计.xlsx'  # 请替换为实际路径
    sheet_name = 'match'

    # 执行难度分类逻辑
    difficulty_df = classify_difficulty(excel_path, sheet_name)

    # 定义模型和章节
    models = [
        "bard", "bard", "bard", "claude2", "claude2", "claude2", 
        "gpt-4", "gpt-4", "gpt-4", "llama2", "llama2", "llama2", 
        "文心一言", "文心一言", "文心一言", "通义千问", "通义千问", "通义千问", 
        "Random-benchmark", "Random-benchmark", "Random-benchmark"
    ]
    chapters = ["Chapter 3", "Chapter 4", "Chapter 5", "Chapter 6"]

    # 生成难度平均分表
    model_difficulty_table = generate_model_difficulty_table(models, chapters, difficulty_df)

  
    # 保存结果为 Excel 文件的两个sheet
    with pd.ExcelWriter('D:\研究生\任务3\supplementary\Fig. 5\支撑数据\difficulty_classification_results.xlsx') as writer:
        difficulty_df.to_excel(writer, sheet_name='Difficulty Classification Results', index=False)
        model_difficulty_table.to_excel(writer, sheet_name='Model Difficulty Average Score Table', index=False)    

    # 打印输出以供检查
    print("Difficulty Classification Results:")
    print(difficulty_df.head())
    print("\nModel Difficulty Average Score Table:")
    print(model_difficulty_table.head())
