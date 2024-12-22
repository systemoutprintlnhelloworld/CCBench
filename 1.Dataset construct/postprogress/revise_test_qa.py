import openpyxl

# 指定新Excel文件路径
file_path = r"D:\研究生\任务3\题库\QA\QA题库(new)_modified.xlsx"

# 打开Excel文件
workbook = openpyxl.load_workbook(file_path)

# 选择第一个工作表
sheet = workbook.active

# 获取工作表的最大行数
max_row = sheet.max_row

# 循环遍历每一行
for row_num in range(1, max_row + 1):
    # 获取第四列和第六列的单元格
    cell_4 = sheet.cell(row=row_num, column=4)
    cell_6 = sheet.cell(row=row_num, column=6)
    
    # 检查第四列是否非空
    if cell_4.value is not None:
        # 将第三列内容修改为第四列内容
        sheet.cell(row=row_num, column=3, value=cell_4.value)
        # 将第五列内容修改为第六列内容
        sheet.cell(row=row_num, column=5, value=cell_6.value)

# 保存修改后的文件
workbook.save(file_path)

# 打印完成提示
print('完成')
