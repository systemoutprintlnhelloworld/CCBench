import openpyxl

# 指定Excel文件路径
file_path = r"D:\研究生\任务3\题库\VQA题库(new)_modified.xlsx"

# 打开Excel文件
workbook = openpyxl.load_workbook(file_path)

# 选择第一个工作表
sheet = workbook.active

# 获取工作表的最大行数
max_row = sheet.max_row

# 循环遍历每一行
for row_num in range(1, max_row + 1):
    # 获取第三列的单元格
    cell_3 = sheet.cell(row=row_num, column=3)
    
    # 检查第二列是否非空
    if cell_3.value is not None:
        # 将第一列内容修改为第二列内容
        sheet.cell(row=row_num, column=1, value=cell_3.value)
        # 将第三列内容修改为'No'
        sheet.cell(row=row_num, column=4, value='No')

# 保存修改后的文件
workbook.save(file_path)

# 打印完成提示
print('完成')
